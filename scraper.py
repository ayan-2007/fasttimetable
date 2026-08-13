"""
FastSchedule - Automated University Timetable

Resilient backend scraper engine.

Pipeline:
  1. Transform the public Google Sheets "view" URL into a direct XLSX export URL.
  2. Download the workbook, read every sheet, and detect the header row using
     fuzzy regex column matching so renamed or moved columns never break parsing.
  3. Normalize every time value into the canonical "HH:MM - HH:MM" format.
  4. Clean the rows and persist them as plain, readable JSON in db/timetable.json.

Fail-safe behavior:
  If any network call or parsing step fails, an alert is logged and the last
  known valid db/timetable.json is left untouched. A corrupt or partial file is
  never written.
"""

import difflib
import json
import os
import re
import sys
import tempfile
from datetime import datetime, timezone

import pandas
import requests
from openpyxl import load_workbook

SHEET_URL = "https://docs.google.com/spreadsheets/d/1fL2TWhPgbPc2d66vm_KywTpdsGBIaBLqlmz4JLPudCw/edit?usp=sharing"

FSM_SHEET_URL = "https://docs.google.com/spreadsheets/d/1AnFQQhv9lu4grESE2ypbDG7E1QOPGgGCRiejem5ocPw/edit?usp=sharing"

SOURCE_URLS = [SHEET_URL, FSM_SHEET_URL]

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(SCRIPT_DIR, "db", "timetable.json")

REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
    )
}
DOWNLOAD_TIMEOUT = 60
MIN_XLSX_PAYLOAD_BYTES = 1024

DAY_ORDER = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

EMPTY_MARKERS = {
    "free",
    "lunch",
    "break",
    "empty",
    "blank",
    "available",
    "no class",
    "noclass",
    "cancelled",
    "canceled",
    "cancel",
    "holiday",
    "n/a",
    "na",
    "nil",
    "none",
    "off",
    "—",
    "-",
    "–",
}

COLUMN_PATTERNS = [
    ("department", re.compile(r"department|dept\.?|program|stream|school|discipline", re.IGNORECASE)),
    ("batch_section", re.compile(r"batch|section|\bgroup\b|division", re.IGNORECASE)),
    ("day", re.compile(r"\bday\b|weekday|\bdate\b|session", re.IGNORECASE)),
    ("time_start", re.compile(r"start(?:\s*time)?|begin(?:\s*time)?|\bfrom\b", re.IGNORECASE)),
    ("time_end", re.compile(r"end(?:\s*time)?|finish(?:\s*time)?|\bto\b", re.IGNORECASE)),
    ("time", re.compile(r"\btime\b|\bslot\b|period|timings?|hours?", re.IGNORECASE)),
    ("course", re.compile(r"course|subject|module|title|offering|paper", re.IGNORECASE)),
    ("instructor", re.compile(r"instructor|teacher|lecturer|faculty|professor|staff", re.IGNORECASE)),
    ("room", re.compile(r"room|venue|classroom|\bhall\b|building|location|block", re.IGNORECASE)),
    ("type", re.compile(r"\btype\b|mode|kind|format|category", re.IGNORECASE)),
]

SHEET_URL_RE = re.compile(r"/spreadsheets/d/([a-zA-Z0-9_-]+)")
TIME_TOKEN_RE = re.compile(
    r"\d{1,2}[:.]\d{1,2}\s*(?:[apAP](?:\.?[mM]\.?)?)?"
    r"|\d{1,2}\s*(?:[apAP](?:\.?[mM]\.?)?)"
)
TIME_LABEL_RE = re.compile(
    r"\d{1,2}[:.]\d{1,2}\s*(?:[apAP](?:\.?[mM]\.?)?)?"
    r"(?:\s*(?:-|–|—|to)\s*\d{1,2}[:.]\d{1,2}\s*(?:[apAP](?:\.?[mM]\.?)?)?)?"
)
AM_RE = re.compile(r"[aA](?:\.?[mM]\.?)?")
PM_RE = re.compile(r"[pP](?:\.?[mM]\.?)?")

DAY_NAME_MAP = {
    "monday": "Mon",
    "tuesday": "Tue",
    "wednesday": "Wed",
    "thursday": "Thu",
    "friday": "Fri",
    "saturday": "Sat",
    "sunday": "Sun",
    "mon": "Mon",
    "tue": "Tue",
    "tues": "Tue",
    "wed": "Wed",
    "thu": "Thu",
    "thur": "Thu",
    "thurs": "Thu",
    "fri": "Fri",
    "sat": "Sat",
    "sun": "Sun",
}

PROGRAM_MAP = {
    "CE": "Computer Engineering",
    "EE": "Electrical Engineering",
    "CS": "Computer Science",
    "SE": "Software Engineering",
    "AI": "Artificial Intelligence",
    "DS": "Data Science",
    "ME": "Mechanical Engineering",
    "BM": "Business Administration",
    "BA": "Business Administration",
    "PE": "Petroleum Engineering",
    "PG": "Physics",
    "PH": "Physics",
}

FSM_PROGRAM_MAP = {
    "BBA": "Bachelor of Business Administration",
    "AF": "Bachelor of Science (Accounting & Finance)",
    "BSAF": "Bachelor of Science (Accounting & Finance)",
    "BSBA": "Bachelor of Science (Business Analytics)",
    "BA": "Bachelor of Science (Business Analytics)",
    "FT": "Bachelor of Science (Financial Technology)",
    "BSFT": "Bachelor of Science (Financial Technology)",
    "MBA": "MBA",
    "MSBA": "MS Business Analytics",
    "MS": "MS Business Analytics",
    "PhD": "PhD",
}

FSM_BATCH_RE = re.compile(
    r"^(?P<prog>BBA|BSBA|BSFT|BSAF|AF|BA|FT|MBA|MSBA|MS|PhD)"
    r"[- ]*(?P<sem>\d{1,2})(?P<sec>[A-E])?"
    r"(?:/(?P<sec2>[A-E]))?"
    r"(?:/(?P<comb>(?:BBA|BSBA|BSFT|BSAF|AF|BA|FT|MBA|MSBA|MS|PhD)[- ]*\d{1,2}[A-E]?))?"
    r"(?P<grp>\d)?$"
)

FSM_EMBED_TIME_RE = re.compile(r"\((\d{1,2}:\d{2})\s*(?:-|–|—)\s*(\d{1,2}:\d{2})\)")

PROGRAM_CODES = "CE|EE|CS|SE|AI|DS|ME|BM|BA|PE|PG|PH|MS"

# Section markers: optional program code + optional semester digit + section letter,
# e.g. "CE-A", "EE-2B", " MS-A", "- A".
MATRIX_SECTION_RE = re.compile(
    r"(?:(\b(?:" + PROGRAM_CODES + r")\b)[ ._\-]*)?(\d)?[ ._\-]*([A-E])$"
)

# Batch/semester detection ---------------------------------------------------
# The "Course Allocation" sheet maps every course to its program + semester
# (e.g. "BS CE 1st Semester" -> Batch BS(CE) 2026). We combine that with the
# color-coded batch blocks in the schedule sheet so every class gets a batch
# identifier like "CE-1A" (CE program, 1st semester, section A).

ALLOC_HEADER_RE = re.compile(r"^\s*BS\s+(\w+)\s+(\d+)(?:st|nd|rd|th)\s+Semester")
ALLOC_SHEET_RE = re.compile(r"allocation|allocated courses", re.IGNORECASE)
MATCH_THRESHOLD = 0.60

COURSE_ABBR = {
    "adv": "advanced",
    "anal": "analysis",
    "analogue": "analog",
    "arch": "architecture",
    "circ": "circuits",
    "com": "communication",
    "comm": "communication",
    "comminity": "community",
    "comp": "complex",
    "dev": "devices",
    "elect": "electronics",
    "eng": "engineering",
    "engg": "engineering",
    "engr": "engineers",
    "fund": "fundamentals",
    "inst": "instrumentation",
    "instru": "instrumentation",
    "intel": "intelligence",
    "inter": "interfacing",
    "mech": "mechanical",
    "mgmt": "management",
    "mp": "microprocessor",
    "net": "network",
    "netwk": "network",
    "netwks": "networks",
    "obj": "object",
    "ocp": "occupational",
    "org": "organization",
    "prog": "programming",
    "struct": "structures",
    "strucures": "structures",
    "sys": "systems",
    "tech": "technical",
    "thermo": "thermodynamics",
    "trans": "transforms",
    "var": "variable",
    "vars": "variables",
}

SECTION_STRIP_RE = re.compile(
    r"\s+(?:(?:" + PROGRAM_CODES + r")\b[ ._\-]*)?(?:\d[ ._\-]*)?[A-E]"
    r"(?:\s*,\s*(?:(?:" + PROGRAM_CODES + r")\b[ ._\-]*)?(?:\d[ ._\-]*)?[A-E])*\s*$"
)

INSTRUCTOR_CUT_RE = re.compile(r"\s+(?:mr\.?|ms\.?|dr\.?|prof\.?)\s+[a-z]", re.IGNORECASE)
TEACHER_CUT_RE = re.compile(r"\s+teacher\s*:", re.IGNORECASE)
TIME_CUT_RE = re.compile(r"\d{1,2}:\d{2}")
PAREN_RE = re.compile(r"\s*\([^)]*\)\s*")


def log_alert(message):
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    print(f"[ALERT] {stamp} {message}", file=sys.stderr)


def get_export_url(url):
    match = SHEET_URL_RE.search(url)
    if not match:
        raise ValueError(f"Could not extract spreadsheet id from URL: {url}")
    spreadsheet_id = match.group(1)
    return f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"


def clean_cell(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float):
            if value == int(value):
                return str(int(value))
            return f"{value:.2f}"
        return str(value)
    text = str(value)
    text = text.replace("\u00a0", " ").replace("\u3000", " ")
    text = re.sub(r"[\r\n\t]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text if text else None


def clean_header(value):
    return (clean_cell(value) or "").lower()


def is_blank(value):
    if value is None:
        return True
    if isinstance(value, (int, float)):
        return False
    text = str(value).strip().lower()
    return not text or text in EMPTY_MARKERS


def _token_to_minutes(token):
    token = token.strip()
    if not token:
        return None
    is_pm = bool(PM_RE.search(token))
    is_am = bool(AM_RE.search(token)) and not is_pm
    match = re.match(r"(\d{1,2})(?:[:.](\d{1,2}))?", token)
    if not match:
        return None
    hour = int(match.group(1))
    minute = int(match.group(2)) if match.group(2) else 0
    if minute >= 60:
        return None
    if is_pm and hour < 12:
        hour += 12
    if is_am and hour == 12:
        hour = 0
    return hour * 60 + minute


def normalize_time(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        text = clean_cell(value)
    else:
        text = str(value)
    if not text:
        return None
    text = re.sub(r"\s+", " ", text).strip()
    tokens = TIME_TOKEN_RE.findall(text)
    minutes = [_token_to_minutes(tok) for tok in tokens]
    minutes = [m for m in minutes if m is not None]
    if len(minutes) == 0:
        return None
    if len(minutes) == 1:
        total = minutes[0]
        return f"{total // 60:02d}:{total % 60:02d}"
    start, end = minutes[0], minutes[-1]
    return f"{start // 60:02d}:{start % 60:02d} - {end // 60:02d}:{end % 60:02d}"


def has_time_label(value):
    text = clean_cell(value)
    if not text:
        return False
    return bool(TIME_LABEL_RE.fullmatch(text))


def first_cell_in_range(row, start, end):
    for col_index in range(start, min(end, len(row) - 1) + 1):
        value = clean_cell(row[col_index])
        if value:
            return value
    return None


def find_matrix_header_row(rows, max_rows=30):
    best_index = -1
    best_score = 0
    for index, row in enumerate(rows[:max_rows]):
        headers = [clean_header(c) for c in row]
        score = 0
        if any(header == "room" for header in headers if header):
            score += 2
        for raw in row:
            if has_time_label(clean_cell(raw)):
                score += 1
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def build_matrix_slots(rows, header_index):
    header_row = rows[header_index]
    merged_spans = {}
    for col_index, raw in enumerate(header_row):
        if has_time_label(clean_cell(raw)):
            merged_spans[col_index] = (col_index, col_index)
    slots = []
    for col_index, raw in enumerate(header_row):
        if col_index not in merged_spans:
            continue
        value = clean_cell(raw)
        label = normalize_time(value) or value
        slots.append({"start": col_index, "end": merged_spans[col_index][1], "label": label})
    slots.sort(key=lambda slot: slot["start"])
    return slots


def expand_slot_spans(worksheet, rows, header_index, slots):
    header_row = rows[header_index]
    for rng in worksheet.merged_cells.ranges:
        if rng.min_row - 1 != header_index:
            continue
        start = rng.min_col - 1
        end = rng.max_col - 1
        for slot in slots:
            if start == slot["start"]:
                slot["end"] = end
                break
    return slots


def find_room_column(rows, header_index):
    for col_index, raw in enumerate(rows[header_index]):
        if clean_header(raw) == "room":
            return col_index
    return 2


def find_day_blocks(rows):
    blocks = []
    current_start = None
    current_day = None
    for index, row in enumerate(rows):
        first = clean_cell(row[0]) if row else None
        day = DAY_NAME_MAP.get(first.lower()) if first else None
        if day:
            if current_start is not None:
                blocks.append((current_start, index - 1, current_day))
            current_start = index
            current_day = day
    if current_start is not None:
        blocks.append((current_start, len(rows) - 1, current_day))
    return blocks


def build_course_rows(worksheet, rows, block_start, block_end, room_col):
    course_rows = set()
    for rng in worksheet.merged_cells.ranges:
        if rng.min_col - 1 != room_col:
            continue
        anchor = rng.min_row - 1
        if block_start <= anchor <= block_end:
            course_rows.add(anchor)
    for index in range(block_start, block_end + 1):
        row = rows[index]
        if room_col < len(row) and clean_cell(row[room_col]) and not has_time_label(clean_cell(row[room_col])):
            course_rows.add(index)
    return course_rows


def extract_section_info(course):
    text = clean_cell(course) or ""
    text = re.sub(r"\s*\([^)]*\)\s*$", "", text)
    match = MATRIX_SECTION_RE.search(text)
    if not match:
        return None, None, None
    program = match.group(1)
    semester = int(match.group(2)) if match.group(2) else None
    section = match.group(3)
    return program, semester, section


def program_in_name(course):
    text = clean_cell(course) or ""
    match = re.search(r"\b(" + PROGRAM_CODES + r")\b", text, re.IGNORECASE)
    return match.group(1).upper() if match else None


def normalize_course(text):
    text = clean_cell(text) or ""
    text = text.lower()
    text = text.replace("&", " and ")
    text = re.sub(r"\ba & digital\b", "analog and digital", text)
    text = re.sub(r"[^a-z0-9.\s/]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    tokens = []
    for token in text.split():
        base = token[:-1] if token.endswith(".") and len(token) > 1 else token
        tokens.append(COURSE_ABBR.get(base, token))
    return " ".join(tokens)


def base_course_name(course):
    text = clean_cell(course) or ""
    text = INSTRUCTOR_CUT_RE.split(text, maxsplit=1)[0]
    text = TEACHER_CUT_RE.split(text, maxsplit=1)[0]
    text = PAREN_RE.sub(" ", text)
    text = TIME_CUT_RE.split(text, maxsplit=1)[0]
    text = SECTION_STRIP_RE.sub("", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def parse_allocation_sheet(worksheet):
    rows = []
    current = None
    for row in worksheet.iter_rows(values_only=True):
        if not row:
            continue
        label = clean_cell(row[0]) if len(row) > 0 else None
        if label:
            match = ALLOC_HEADER_RE.match(label)
            if match:
                current = (match.group(1).upper(), int(match.group(2)))
                continue
        if current:
            code = clean_cell(row[2]) if len(row) > 2 else None
            name = clean_cell(row[3]) if len(row) > 3 else None
            if code and name:
                normalized = normalize_course(name)
                if normalized:
                    rows.append((current[0], current[1], normalized))
    return rows


def match_allocation(name, program_hint, alloc_rows):
    base = normalize_course(base_course_name(name))
    if not base:
        return None, 0.0, False
    best = None
    best_ratio = 0.0
    tied = False
    for program, semester, alloc_name in alloc_rows:
        if program_hint and program != program_hint:
            continue
        ratio = difflib.SequenceMatcher(None, base, alloc_name).ratio()
        if ratio > best_ratio + 1e-6:
            best_ratio = ratio
            best = (program, semester)
            tied = False
        elif abs(ratio - best_ratio) <= 1e-6 and ratio > 0:
            tied = True
    if best_ratio < MATCH_THRESHOLD:
        return None, best_ratio, False
    return best, best_ratio, tied


def cell_fill_key(worksheet, row_index, col_index):
    try:
        cell = worksheet.cell(row_index + 1, col_index + 1)
    except (IndexError, ValueError):
        return None
    fill = cell.fill
    if fill is None or fill.fill_type is None:
        return None
    rgb = fill.fgColor.rgb if fill.fgColor else None
    if not rgb or str(rgb) == "00000000":
        return None
    return str(rgb)


def first_cell_anchor(row, start, end):
    for col_index in range(start, min(end, len(row) - 1) + 1):
        if clean_cell(row[col_index]):
            return col_index
    return None


def compose_batch(program, semester, section):
    if semester is not None:
        prefix = f"{program}-{semester}" if program else str(semester)
    elif program:
        prefix = program
    else:
        prefix = ""
    if not section or section == "All":
        return f"{prefix} (All)" if prefix else "All"
    if semester is not None:
        return f"{prefix}{section}"
    return f"{prefix}-{section}" if prefix else section


def enrich_batch_info(entries, alloc_rows):
    for entry in entries:
        hint = entry.get("_prog_hint")
        matched, ratio, tied = match_allocation(entry.get("course"), hint, alloc_rows)
        entry["_matched_program"] = matched[0] if matched else None
        entry["_matched_semester"] = matched[1] if matched else None
        entry["_tied"] = tied and matched is not None

    votes = {}
    for entry in entries:
        program = entry.get("_matched_program")
        semester = entry.get("_matched_semester")
        color = entry.get("_color")
        if program is not None and semester is not None and color:
            key = (program, semester)
            votes.setdefault(color, {}).setdefault(key, 0)
            votes[color][key] += 1
    color_assign = {}
    for color, counter in votes.items():
        total = sum(counter.values())
        top, count = max(counter.items(), key=lambda item: item[1])
        if count > total / 2.0:
            color_assign[color] = top

    for entry in entries:
        name = entry.get("course")
        hint = entry.get("_prog_hint")
        semester_digit = entry.get("_sem_digit")
        section = entry.get("_section")
        matched_program = entry.get("_matched_program")
        matched_semester = entry.get("_matched_semester")
        tied = entry.get("_tied", False)

        program = matched_program if matched_program else hint
        semester = matched_semester
        if semester_digit is not None:
            program = hint or program
            semester = semester_digit
        elif not program or semester is None or tied:
            color = entry.get("_color")
            if color in color_assign:
                program, semester = color_assign[color]

        if program or semester is not None or section:
            if program:
                entry["department"] = PROGRAM_MAP.get(program, "School of Engineering")
            entry["semester"] = semester
            entry["section"] = section or "All"
            entry["batch_section"] = compose_batch(program, semester, section)
        else:
            entry["semester"] = None
            entry["section"] = entry.get("batch_section") or "All"

        for key in ("_color", "_prog_hint", "_sem_digit", "_section", "_matched_program", "_matched_semester", "_tied"):
            entry.pop(key, None)
    return entries


def build_matrix_entry(day, time_label, course, instructor, room, color):
    program, semester_digit, section = extract_section_info(course)
    return {
        "department": PROGRAM_MAP.get(program, "School of Engineering"),
        "batch_section": compose_batch(program, semester_digit, section),
        "day": day,
        "time": time_label,
        "course": course,
        "instructor": instructor or "Not assigned",
        "room": room or "Not specified",
        "type": classify_type(None, course),
        "_color": color,
        "_prog_hint": program,
        "_sem_digit": semester_digit,
        "_section": section,
    }


def parse_matrix(worksheet, rows, header_index):
    slots = build_matrix_slots(rows, header_index)
    if len(slots) < 2:
        return []
    slots = expand_slot_spans(worksheet, rows, header_index, slots)
    room_col = find_room_column(rows, header_index)
    entries = []
    for block_start, block_end, day in find_day_blocks(rows):
        slot_labels = [slot["label"] for slot in slots]
        course_rows = build_course_rows(worksheet, rows, block_start, block_end, room_col)
        for row_index in range(block_start, block_end + 1):
            row = rows[row_index]
            is_subheader = False
            for index, slot in enumerate(slots):
                value = first_cell_in_range(row, slot["start"], slot["end"])
                if has_time_label(value):
                    slot_labels[index] = normalize_time(value) or value
                    is_subheader = True
            if is_subheader:
                continue
            if row_index not in course_rows:
                continue
            room_value = clean_cell(row[room_col]) if room_col < len(row) else None
            for index, slot in enumerate(slots):
                course = first_cell_in_range(row, slot["start"], slot["end"])
                if not course or has_time_label(course):
                    continue
                lowered = course.lower()
                if re.search(r"reserv|resrv|resev", lowered):
                    continue
                if re.search(r"faculty\s+meeting", lowered):
                    continue
                instructor = None
                if row_index + 1 < len(rows):
                    instructor = first_cell_in_range(rows[row_index + 1], slot["start"], slot["end"])
                anchor = first_cell_anchor(row, slot["start"], slot["end"])
                color = cell_fill_key(worksheet, row_index, anchor) if anchor is not None else None
                entry = build_matrix_entry(day, slot_labels[index], course, instructor, room_value, color)
                entry["_prog_hint"] = program_in_name(course) or entry["_prog_hint"]
                entries.append(entry)
    return entries


FSM_BATCH_TOKEN_RE = re.compile(
    r"^(?P<prog>BBA|BSBA|BSFT|BSAF|AF|BA|FT|MBA|MSBA|MS|PhD)[- ]*(?P<sem>\d{1,2})(?P<sec>[A-E])?(?P<grp>\d)?$"
)


def is_fsm_subheader(row):
    hits = 0
    for col_index in range(3, 64):
        if col_index < len(row) and has_time_label(clean_cell(row[col_index])):
            hits += 1
    return hits >= 4


def fsm_logical_cells(worksheet, row_index):
    r1 = row_index + 1
    merged = []
    for rng in worksheet.merged_cells.ranges:
        if rng.min_row == r1:
            merged.append((rng.min_col - 1, rng.max_col - 1))
    used = set()
    for start, end in merged:
        for col_index in range(start, end + 1):
            used.add(col_index)
    cells = []
    for start, end in sorted(merged):
        value = clean_cell(worksheet.cell(r1, start + 1).value)
        cells.append((start, end, value))
    for col_index in range(0, worksheet.max_column):
        if col_index in used:
            continue
        value = clean_cell(worksheet.cell(r1, col_index + 1).value)
        if value:
            cells.append((col_index, col_index, value))
    cells.sort()
    return cells


def format_fsm_batch(code):
    parts = [part.strip() for part in str(code).split("/") if part.strip()]
    out = []
    prev = None
    for part in parts:
        if re.fullmatch(r"[A-E]", part) and prev is not None:
            out[-1] = out[-1] + "/" + part
            continue
        token = FSM_BATCH_TOKEN_RE.fullmatch(part)
        if not token:
            return None
        label = f"{token.group('prog')}-{int(token.group('sem'))}".replace(".0", "")
        label = f"{token.group('prog')}-{int(token.group('sem'))}"
        if token.group("sec"):
            label += token.group("sec")
        if token.group("grp"):
            label = f"{label} (G{token.group('grp')})"
        out.append(label)
        prev = token
    return "/".join(out) if out else None


def parse_fsm_batch(code):
    token = FSM_BATCH_TOKEN_RE.fullmatch(str(code))
    if not token:
        return None
    prog = token.group("prog")
    semester = int(token.group("sem"))
    section = token.group("sec") or "All"
    return prog, semester, section


def fsm_time_for(course, start_col, header_row):
    text = clean_cell(course) or ""
    embedded = FSM_EMBED_TIME_RE.search(text)
    if embedded:
        return normalize_time(f"{embedded.group(1)}-{embedded.group(2)}")
    best = None
    for col_index, raw in enumerate(header_row):
        if col_index >= 3 and col_index <= start_col and has_time_label(clean_cell(raw)):
            best = col_index
    if best is not None and best < len(header_row):
        return normalize_time(header_row[best])
    return None


def parse_fsm_matrix(worksheet, rows, header_index):
    header_row = rows[header_index]
    entries = []
    for block_start, block_end, day in find_day_blocks(rows):
        for row_index in range(block_start, block_end + 1):
            row = rows[row_index]
            if is_fsm_subheader(row):
                continue
            room = row[2] if len(row) > 2 else None
            if not room:
                continue
            cells = [cell for cell in fsm_logical_cells(worksheet, row_index) if cell[1] >= 3]
            pending_course = None
            pending_slot = None
            skipped_fillers = ("MS", "CS")
            for start, end, value in cells:
                if not value:
                    continue
                if value in skipped_fillers:
                    continue
                if has_time_label(value):
                    continue
                if format_fsm_batch(value) is not None:
                    if pending_course is not None:
                        batch = parse_fsm_batch(value)
                        if batch is not None:
                            prog, semester, section = batch
                            time_label = fsm_time_for(pending_course, pending_slot, header_row)
                            department = FSM_PROGRAM_MAP.get(prog, "FAST School of Management")
                            entries.append(
                                {
                                    "department": department,
                                    "batch_section": format_fsm_batch(value),
                                    "day": day,
                                    "time": time_label,
                                    "course": base_course_name(pending_course),
                                    "instructor": "Not assigned",
                                    "room": room,
                                    "type": classify_type(None, pending_course),
                                    "semester": semester,
                                    "section": section,
                                    "school": "Management",
                                }
                            )
                        pending_course = None
                        pending_slot = None
                    continue
                if pending_course is not None:
                    pending_course = None
                    pending_slot = None
                pending_course = value
                pending_slot = start
    return entries


def resolve_merged_cells(worksheet, rows):
    if not worksheet.merged_cells.ranges:
        return rows
    for rng in worksheet.merged_cells.ranges:
        top = rng.min_row - 1
        left = rng.min_col - 1
        if top >= len(rows) or left >= len(rows[top]):
            continue
        anchor = rows[top][left]
        if anchor is None:
            continue
        for r in range(top, min(rng.max_row, len(rows))):
            for c in range(left, min(rng.max_col, len(rows[r]))):
                if rows[r][c] is None:
                    rows[r][c] = anchor
    return rows


def find_header_row(rows, max_rows=25):
    best_index = -1
    best_score = -1
    for index, row in enumerate(rows[:max_rows]):
        headers = [clean_header(c) for c in row]
        score = 0
        for _field, pattern in COLUMN_PATTERNS:
            if any(pattern.search(header) for header in headers if header):
                score += 1
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def map_columns(header_row):
    mapping = {}
    for col_index, raw in enumerate(header_row):
        header = clean_header(raw)
        if not header:
            continue
        for field, pattern in COLUMN_PATTERNS:
            if pattern.search(header):
                mapping.setdefault(field, []).append(col_index)
                break
    return mapping


def _first_cell(field, mapping, row):
    for col_index in mapping.get(field, []):
        if col_index >= len(row):
            continue
        value = clean_cell(row[col_index])
        if value and not is_blank(value):
            return value
    return None


def resolve_time(field_single, mapping, row):
    single = _first_cell(field_single, mapping, row)
    if single:
        return normalize_time(single)
    start = _first_cell("time_start", mapping, row)
    end = _first_cell("time_end", mapping, row)
    if start or end:
        combined = " - ".join(part for part in (start, end) if part)
        if combined:
            return normalize_time(combined)
    return None


def classify_type(type_raw, course_raw):
    raw = clean_cell(type_raw)
    if raw:
        text = raw.lower()
    else:
        text = (clean_cell(course_raw) or "").lower()
    if re.search(r"\blab\b|laboratory|laborator", text):
        return "Lab"
    return "Lecture"


def is_cancelled_entry(values):
    joined = " | ".join(str(v) for v in values if v is not None).lower()
    return any(marker in joined for marker in ("cancelled", "canceled", "cancell"))


def parse_flat(rows, header_index):
    mapping = map_columns(rows[header_index])
    required_fields = ("day", "course", "time")
    if not all(mapping.get(field) for field in required_fields):
        return []

    entries = []
    for row in rows[header_index + 1 :]:
        day = _first_cell("day", mapping, row)
        course = _first_cell("course", mapping, row)
        if not day or not course:
            continue
        if is_blank(course):
            continue
        if is_cancelled_entry(row):
            continue
        time_value = resolve_time("time", mapping, row)
        if not time_value:
            continue
        department = _first_cell("department", mapping, row)
        batch_section = _first_cell("batch_section", mapping, row)
        instructor = _first_cell("instructor", mapping, row)
        room = _first_cell("room", mapping, row)
        type_raw = _first_cell("type", mapping, row)
        entry_type = classify_type(type_raw, course)

        entry = {
            "department": department or "General",
            "batch_section": batch_section or "All",
            "day": str(day).strip(),
            "time": time_value,
            "course": course,
            "instructor": instructor or "Not assigned",
            "room": room or "Not specified",
            "type": entry_type,
        }
        entries.append(entry)
    return entries


def parse_sheet(worksheet, sheet_name, rows):
    matrix_header = find_matrix_header_row(rows)
    if matrix_header >= 0:
        entries = parse_matrix(worksheet, rows, matrix_header)
        if entries:
            return entries
    flat_header = find_header_row(rows)
    if flat_header < 0:
        return []
    merged_rows = resolve_merged_cells(worksheet, rows)
    return parse_flat(merged_rows, flat_header)


def read_workbook_rows(worksheet, workbook_path, sheet_name):
    frame = pandas.read_excel(
        workbook_path,
        sheet_name=sheet_name,
        header=None,
        dtype=str,
        engine="openpyxl",
    )
    rows = frame.astype(object).where(pandas.notna(frame), None).values.tolist()
    return [[clean_cell(cell) for cell in row] for row in rows]


def scrape_workbook(payload, school="Engineering"):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(payload)
        tmp_path = tmp.name
    try:
        workbook = load_workbook(tmp_path, read_only=False, data_only=True)
        entries = []
        sheet_stats = {}
        if school == "Management":
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                if not re.search(r"timetable", sheet_name, re.IGNORECASE):
                    continue
                rows = read_workbook_rows(worksheet, tmp_path, sheet_name)
                if not rows:
                    continue
                matrix_header = find_matrix_header_row(rows)
                if matrix_header < 0:
                    continue
                sheet_entries = parse_fsm_matrix(worksheet, rows, matrix_header)
                sheet_stats[sheet_name] = len(sheet_entries)
                entries.extend(sheet_entries)
        else:
            alloc_rows = []
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                if ALLOC_SHEET_RE.search(sheet_name):
                    alloc_rows.extend(parse_allocation_sheet(worksheet))
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                if ALLOC_SHEET_RE.search(sheet_name):
                    continue
                rows = read_workbook_rows(worksheet, tmp_path, sheet_name)
                if not rows:
                    continue
                sheet_entries = parse_sheet(worksheet, sheet_name, rows)
                sheet_stats[sheet_name] = len(sheet_entries)
                entries.extend(sheet_entries)
            entries = enrich_batch_info(entries, alloc_rows)
            for entry in entries:
                entry["school"] = "Engineering"
            print(f"[INFO] Allocation rows: {len(alloc_rows)}")
        workbook.close()
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

    entries.sort(key=lambda item: (
        DAY_ORDER.index(item["day"]) if item["day"] in DAY_ORDER else len(DAY_ORDER),
        item["time"],
        item["batch_section"].lower(),
    ))
    print(f"[INFO] Sheets parsed: {sheet_stats}")
    return entries


def write_db(entries):
    payload = {
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S%z"),
        "source": ", ".join(SOURCE_URLS),
        "count": len(entries),
        "entries": entries,
    }
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    temp_path = DB_PATH + ".tmp"
    with open(temp_path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2, ensure_ascii=False)
    os.replace(temp_path, DB_PATH)
    print(f"[INFO] Wrote {len(entries)} entries to {DB_PATH}")


def download_workbook(url):
    export_url = get_export_url(url)
    print(f"[INFO] Export URL: {export_url}")
    response = requests.get(export_url, timeout=DOWNLOAD_TIMEOUT, headers=REQUEST_HEADERS)
    response.raise_for_status()
    if len(response.content) < MIN_XLSX_PAYLOAD_BYTES:
        raise RuntimeError(
            f"Export payload unexpectedly small ({len(response.content)} bytes); "
            "the sheet may require sign-in or returned an error page."
        )
    if response.content[:2] not in (b"PK", b"<?") and b"<html" in response.content[:512].lower():
        raise RuntimeError("Export returned an HTML page instead of a workbook.")
    return response.content


def scrape():
    entries = []
    for source_url in SOURCE_URLS:
        payload = download_workbook(source_url)
        school = "Management" if source_url == FSM_SHEET_URL else "Engineering"
        entries.extend(scrape_workbook(payload, school=school))
    return entries


def main():
    try:
        entries = scrape()
    except Exception as exc:
        log_alert(f"Scrape failed: {exc}")
        log_alert("Preserving the last known valid db/timetable.json untouched.")
        return 1
    try:
        write_db(entries)
    except Exception as exc:
        log_alert(f"Could not write timetable database: {exc}")
        log_alert("Preserving the last known valid db/timetable.json untouched.")
        return 1
    print("[INFO] Timetable update completed successfully.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
